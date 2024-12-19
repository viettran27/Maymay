import pandas as pd
from fastapi import APIRouter, Response, HTTPException, UploadFile, File, Depends, Form
from openpyxl import Workbook
from io import BytesIO
from app.helper.table import import_to_sql
from sqlalchemy import VARCHAR, NVARCHAR, INTEGER, DATE, text
from app.db.base import engine_1, get_db_1
from sqlalchemy.orm import Session
from app.helper.utils import get_days_in_month
from openpyxl.styles import Border, Side, Alignment
from datetime import datetime, timedelta

router = APIRouter()

def get_data(fac, status, location, month, year, db):
    try:
        res = db.execute(text(f"SELECT * FROM MAYTON WHERE MONTH(Ngay) = {month} AND YEAR(Ngay) = {year} AND Nha_may = '{fac}' AND Trang_thai = N'{status}'{f" AND Chi_tiet = N'{location}'" if location != "null" else ''} ORDER BY Loai_may, Ngay")).fetchall()
        data = [dict(row._mapping) for row in res]
        days = get_days_in_month(month, year)

        new_data = {}
        for row in data:
            loai_may = row["Loai_may"]
            ngay = row["Ngay"].day
            so_luong = row["So_luong"]
            
            if loai_may not in new_data:
                new_data[loai_may] = [0]*days
            
            new_data[loai_may][ngay-1] = so_luong
        
        return new_data
    except Exception as e:
        print(e)
        HTTPException(status_code=500, detail="Internal server error")

@router.get("/export")
async def get_excel(month: int, year: int, fac: str, status: str, db: Session = Depends(get_db_1)):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        days = get_days_in_month(month, year) 
        
        header1 = ["Loại máy"]
        header2 = [f"{i}/{month}" for i in range(1, days + 1)]
        headers = header1 + header2
        ws.append(headers)

        machine = [
            "SN",
            "SN (use auto knife)",
            "OL",
            "OL Top feeder - Máy cổ nhỏ",
            "OL (hide seam inside)",
            "FL (hemming)",
            "FL binding",
            "FL Special (many needles)",
            "FL auto cut",
            "Bartack (BTK)",
            "SN 2K",
            "BTH",
            "LBH (thùa khuy)"
        ]

        data = get_data(fac, status, month, year, db)
        days = get_days_in_month(month, year)
        row = 2
        for i in range(len(machine)):
            ws.cell(row=row, column=1, value=machine[i])
            
            if len(data) != 0:
                for j in range(days):
                    ws.cell(row=row, column=j+2, value=data[machine[i]][j])    

            row += 1

        thin_border = Border(left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"))

        for row in ws.iter_rows(min_row=1, max_row=len(machine) + 1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in range(1, len(machine) + 2):
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = max(adjusted_width,7)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        current_time = datetime.now().strftime("%H_%M_%S")
        return Response(
            buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=Mayton_{0}_{1}_{2}_{3}_{4}.xlsx".format(fac, status, month, year, current_time)},
        )
    except:
        HTTPException(status_code=500, detail="Internal server error")

@router.post("/NT1")
async def post_excel(date_from: datetime = Form(...), date_to: datetime = Form(...), file: UploadFile = File(...)):
    try:
        df = pd.read_excel(file.file, sheet_name='Tổng hợp', skiprows=4, header=[0, 1])
        data = df.to_dict('records')
        
        main = []

        status = ["Máy cho mượn", "Máy mượn", "Máy thuê", "Xuất bán thanh lý"]
        key_pair = {
            "Máy cho mượn": "Cho mượn",
            "Máy mượn": "Mượn",
            "Máy thuê": "Thuê",
            "Xuất bán thanh lý": "Thanh lý"
        }

        machine_data = {}
        for row in data:
            type = row[('Loại máy', 'Unnamed: 1_level_1')]
            if not pd.isna(type):
                if not machine_data.get(type):
                    machine_data[type] = {}

                for key, value in row.items():                    
                    if key[0].strip() in status:
                        stt = key[0].strip()
                        location = key[1].strip()

                        if not machine_data[type].get(stt):
                            machine_data[type][stt] = {}

                        if not machine_data[type][stt].get(location):
                            machine_data[type][stt][location] = 0

                        machine_data[type][stt][location] += 0 if pd.isna(value) else value 
            else:
                continue

        for loai_may, value in machine_data.items():
            for stt, value_stt in value.items():
                for location, total in value_stt.items():
                    current_date = date_from
                    while current_date <= date_to:
                        row_data = {}
                        row_data["Loai_may"] = loai_may
                        row_data["Trang_thai"] = key_pair.get(stt)
                        row_data["Ngay"] = current_date
                        row_data["Nha_may"] = 'NT1'
                        row_data["So_luong"] = total
                        row_data["Chi_tiet"] = location if 'Unnamed' not in location else None
                        current_date += timedelta(days=1)
                        main.append(row_data)

        dtype = {
            "Loai_may": NVARCHAR(200),
            "Trang_thai": NVARCHAR(20),
            "Ngay": DATE,
            "Nha_may": VARCHAR(3),   
            "So_luong": INTEGER,
            "Chi_tiet": NVARCHAR(100)
        }
        frame = pd.DataFrame(main)
        import_to_sql(frame, "MAYTON", dtype, engine_1)

        return {"message": "Cập nhật dữ liệu thành công!"}
    except Exception as e:
        print(e)
        HTTPException(status_code=500, detail="Internal server error")

@router.post("/NT2")
async def post_excel2(date_from: datetime = Form(...), date_to: datetime = Form(...), file: UploadFile = File(...)):
    try:
        df = pd.read_excel(file.file, sheet_name=0, skiprows=7, header=[0, 1])
        data = df.to_dict('records')
        main = []

        status = ["Máy cho Hải Phòng mượn", "Máy mượn Hải Phòng", "Máy thuê"]
        key_pair = {
            "Máy cho Hải Phòng mượn": "Cho mượn",
            "Máy mượn Hải Phòng": "Mượn",
            "Máy thuê": "Thuê"
        }

        machine_data = {}

        for row in data:
            type = row[('Loại máy', 'Unnamed: 2_level_1')]
            if not pd.isna(type):
                if not machine_data.get(type):
                    machine_data[type] = {}

                for key, value in row.items():                    
                    if key[0].strip() in status:
                        stt = key[0].strip()
                        location = key[1].strip()

                        if not machine_data[type].get(stt):
                            machine_data[type][stt] = {}

                        if not machine_data[type][stt].get(location):
                            machine_data[type][stt][location] = 0

                        machine_data[type][stt][location] += 0 if pd.isna(value) else value 
            else:
                continue

        for loai_may, value in machine_data.items():
            for stt, value_stt in value.items():
                for location, total in value_stt.items():
                    current_date = date_from
                    while current_date <= date_to:
                        row_data = {}
                        row_data["Loai_may"] = loai_may
                        row_data["Trang_thai"] = key_pair.get(stt)
                        row_data["Ngay"] = current_date
                        row_data["Nha_may"] = 'NT2'
                        row_data["So_luong"] = total
                        row_data["Chi_tiet"] = location if 'Unnamed' not in location else "Hải Phòng"
                        current_date += timedelta(days=1)
                        main.append(row_data)

        dtype = {
            "Loai_may": NVARCHAR(200),
            "Trang_thai": NVARCHAR(20),
            "Ngay": DATE,
            "Nha_may": VARCHAR(3),   
            "So_luong": INTEGER,
            "Chi_tiet": NVARCHAR(100)
        }
        frame = pd.DataFrame(main)
        import_to_sql(frame, "MAYTON", dtype, engine_1)

        return {"message": "Cập nhật dữ liệu thành công!"}
    except Exception as e:
        print(e)
        HTTPException(status_code=500, detail="Internal server error")

@router.get("/machines")
async def get_machines(db: Session = Depends(get_db_1)):
    try:
        query = text("SELECT DISTINCT Loai_may FROM MAYTON")
        result = db.execute(query).fetchall()
        data = [row[0] for row in result]
        return data
    except Exception as e:
        print(e)
        return []
    

@router.get("/")
async def get_mayton(db: Session = Depends(get_db_1), month: int = None, year: int = None, location: str = None, fac: str = "NT1", status: str = "1"):
    try:
        data = get_data(fac, status, location, month, year, db)
        return {"data": data}
    except:
        HTTPException(status_code=500, detail="Internal server error")