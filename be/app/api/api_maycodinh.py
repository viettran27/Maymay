import pandas as pd
from fastapi import APIRouter, Response, HTTPException, UploadFile, File, Depends, Form
from openpyxl import Workbook
from io import BytesIO
from app.helper.table import import_to_sql
from sqlalchemy import VARCHAR, NVARCHAR, INTEGER, DATE, text
from app.db.base import engine_1, get_db_1
from sqlalchemy.orm import Session
from app.helper.utils import get_days_in_month
from openpyxl.styles import Border, Side, Alignment
from datetime import datetime
import math

router = APIRouter()

REGION = {
    "Tuyen_dung": "Tuyển dụng",
    "Thay_layout": "Thay Layout",
    "Mau": "Mẫu",
    "Tat_ca": "Tất cả"
}

def get_data(fac, region, month, year, db: Session = Depends(get_db_1)):
    try:
        if not year:
            HTTPException(status_code=400, detail="Xin hãy nhập năm!")
        
        query = f"SELECT * FROM MAYCODINH WHERE Nha_may = '{fac}' AND MONTH(Ngay) = {month} AND YEAR(Ngay) = {year} ORDER BY Khu_vuc, Loai_may, NGAY"
        if region != "Tat_ca":
            query = f"SELECT * FROM MAYCODINH WHERE Nha_may = '{fac}' AND MONTH(Ngay) = {month} AND YEAR(Ngay) = {year} AND Khu_vuc = N'{REGION.get(region)}' ORDER BY Khu_vuc, Loai_may, NGAY"

        res = db.execute(text(query)).fetchall()
        data = [dict(row._mapping) for row in res]  

        days = get_days_in_month(month, year) 

        new_data = {}
        for row in data:
            khu_vuc = row["Khu_vuc"]
            loai_may = row["Loai_may"]
            ngay = row["Ngay"].day
            so_luong = row["So_luong"]

            if khu_vuc not in new_data:
                new_data[khu_vuc] = {loai_may: [0]*days}
            
            if loai_may not in new_data[khu_vuc]:
                new_data[khu_vuc][loai_may] = [0]*days
            
            new_data[khu_vuc][loai_may][ngay-1] = so_luong

        return new_data
    except Exception as e:
        print(e)
        HTTPException(status_code=500, detail="Internal server error")
    

@router.get("/export")
async def get_excel(month: int, year: int, fac: str, region: str, db: Session = Depends(get_db_1)):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        days = get_days_in_month(month, year) 
        
        header1 = ["Khu vực", "Loại máy"]
        header2 = [f"{i}/{month}" for i in range(1, days + 1)]
        headers = header1 + header2
        ws.append(headers)

        position = ["Tuyển dụng", "Thay Layout", "Mẫu"]
        if region != "Tat_ca":
            position = [REGION.get(region)]

        machine = [
            "SN",
            "SN (use auto knife)",
            "OL",
            "OL Top feeder - Máy cổ nhỏ",
            "OL (hide seam inside)",
            "FL (hemming)",
            "FL binding",
            "FL Special (many needles)",
            "FL auto cut",
            "Bartack (BTK)",
            "SN 2K",
            "BTH",
            "LBH (thùa khuy)"
        ]

        row = 2

        data = get_data(fac, region, month, year, db)
        days = get_days_in_month(month, year)

        for i in range(len(position)):
            for j in range(len(machine)):
                position_name = position[i]
                machine_name = machine[j]

                ws.cell(row=row, column=1, value=position_name)
                ws.cell(row=row, column=2, value=machine_name)

                if len(data) != 0:
                    quantity = data[position_name][machine_name]
                    for k in range(days):
                        ws.cell(row=row, column=k+3, value=quantity[k])

                row += 1

            
        start_row = 2
        for i in range(len(position)):
            ws.merge_cells(start_row=start_row, end_row=start_row+len(machine) - 1, start_column=1, end_column=1)
            start_row += len(machine)

        thin_border = Border(left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"))

        for row in ws.iter_rows(min_row=1, max_row=len(position) * len(machine) + 1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in range(2, len(position) * len(machine) + 2):
            ws.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='center')

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = max(adjusted_width,7)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        current_time = datetime.now().strftime("%H_%M_%S")
        return Response(
            buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={f"Content-Disposition": "attachment; filename=Maycodinh_{0}_{1}_{2}_{3}_{4}.xlsx".format(fac, region, month, year, current_time)}
        )
    except Exception as e:
        # print(e)
        HTTPException(status_code=500, detail="Internal server error")

@router.post("/")
async def post_excel(file: UploadFile = File(...), month: int = Form(...), year: int = Form(...), fac: str = Form(...)):
    try:
        if not year:
            HTTPException(status_code=400, detail="Xin hãy nhập năm!")

        df = pd.read_excel(file.file)
        df["Khu vực"].ffill(inplace=True)
        
        data = df.to_dict('records')
        new_data = []
        for row in data:
            for col in df.columns[2:]:
                day = col
                if type(col) == int:
                    day = str(col)
                new_data.append({"Khu_vuc": row["Khu vực"].strip(), "Nha_may": fac, "Loai_may": row["Loại máy"], "Ngay": datetime(int(year), int(month), int(day.split("/")[0])), "So_luong": row[col] if not math.isnan(row[col]) else 0})

        dtype = {
            "Khu_vuc": NVARCHAR(100),
            "Nha_may": VARCHAR(3),
            "Loai_may": NVARCHAR(200),
            "Ngay": DATE,
            "So_luong": INTEGER
        }

        df_new = pd.DataFrame(new_data)
        import_to_sql(df_new, "MAYCODINH", dtype, engine_1)

        return {"message": "Cập nhật dữ liệu thành công!"}
    except:
        HTTPException(status_code=500, detail="Internal server error")

@router.get("/")
async def get_fs(db: Session = Depends(get_db_1), region: str = None, month: int = None, year: int = None, fac: str = "NT1"):
    try:
        data = get_data(fac, region, month, year, db)
        return {"data": data}
    except:
        HTTPException(status_code=500, detail="Internal server error")