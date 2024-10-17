import pandas as pd
from fastapi import APIRouter, Response, HTTPException, UploadFile, File, Depends, Body
from openpyxl import Workbook
from io import BytesIO
from openpyxl.utils import get_column_letter
from app.helper.table import import_to_sql
from sqlalchemy import VARCHAR, NVARCHAR, INTEGER, text
from app.db.base import engine_1, get_db_1
from sqlalchemy.orm import Session
from datetime import datetime

router = APIRouter()

def get_data(db: Session = Depends(get_db_1), style: str = ""):
    try:
        res = db.execute(text(f"SELECT * FROM FS WHERE STYLE LIKE '%{style}%'")).fetchall()
        data = [dict(row._mapping) for row in res]

        new_data = {}
        for row in data:
            style = row["Style"]
            if style not in new_data:
                new_data[style] = {}
            new_data[style][row["Loai_may"]] = row["So_luong"]
        
        data = [{"Style": style, **loai_may} for style, loai_may in new_data.items()]
        return data
    except:
        HTTPException(status_code=500, detail="Internal server error")

@router.get("/export")
async def get_excel(db: Session = Depends(get_db_1)):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
            
        headers = [
            "Style",
            "SN",
            "SN (use auto knife)",
            "OL",
            "OL Top feeder - Máy cổ nhỏ",
            "OL (hide seam inside)",
            "FL (hemming)",
            "FL binding",
            "FL Special (many needles)",
            "FL auto cut",
            "Bartack (BTK)",
            "SN 2K",
            "BTH",
            "LBH (thùa khuy)"
        ]
        
        ws.append(headers)
        data = get_data(db)

        for row in data:
            ws.append([row.get(header, "") for header in headers])

        for i, column in enumerate(headers, 1):
            col_letter = get_column_letter(i)
            max_length = len(column) + 5
            ws.column_dimensions[col_letter].width = max_length
            
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        current_time = datetime.now().strftime("%H_%M_%S")
        return Response(
            buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=FS_{0}.xlsx".format(current_time)},
        )
    except:
        HTTPException(status_code=500, detail="Internal server error")

@router.post("/")
async def post_excel(file: UploadFile = File(...)):
    try:
        df = pd.read_excel(file.file)
        data = df.to_dict('records')

        new_data = []
        for row in data:
            for key, value in row.items():
                if key != "Style":
                    new_data.append({"Style": row["Style"], "Loai_may": key, "So_luong": value})
        
        df_new = pd.DataFrame(new_data)
        dtype = {
            "Style": VARCHAR(20),
            "Loai_may": NVARCHAR(200),
            "So_luong": INTEGER
        }

        import_to_sql(df_new, "FS", dtype, engine_1) 

        return {"message": "Thêm dữ liệu thành công!"}
    except:
        HTTPException(status_code=500, detail="Internal server error")

@router.get("/")
async def get_fs(style: str = "", db: Session = Depends(get_db_1)):
    try:
        data = get_data(db, style)
        return {"data": data}
    except:
        HTTPException(status_code=500, detail="Internal server error")

@router.post("/style")
async def post_style(data: dict = Body(...)):
    try:
        new_data = []
        for key, value in data.items():
            if key != "Style":
                new_data.append({"Style": data["Style"], "Loai_may": key, "So_luong": int(value)})
        
        dtype = {
            "Style": VARCHAR(20),
            "Loai_may": NVARCHAR(200),
            "So_luong": INTEGER
        }
        df_new = pd.DataFrame(new_data)
        import_to_sql(df_new, "FS", dtype, engine_1)
        
        return {"message": "Cập nhật dữ liệu thành công!"}
    except:
        HTTPException(status_code=500, detail="Internal server error")


@router.get("/styles_not_have")
async def get_styles_not_have(db: Session = Depends(get_db_1)):
    try:
        res = db.execute(text("SELECT * FROM DS_STYLE_CHUA_CO_FS WHERE STYLE IS NOT NULL")).fetchall()
        data = [dict(row._mapping)["STYLE"] for row in res]
        return {"data": data}
    except:
        HTTPException(status_code=500, detail="Internal server error")